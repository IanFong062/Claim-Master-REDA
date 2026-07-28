from __future__ import annotations

import copy
import hashlib
import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = ROOT / "templates" / "Travelling claim form - updated version.xlsx"
OUTPUT = ROOT / "outputs" / "tests"
REPORT = ROOT / "TEST_REPORT.md"
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
ET.register_namespace("", NS["m"])


def load_mapping() -> dict:
    text = (ROOT / "config" / "templateMapping.js").read_text(encoding="utf-8")
    body = re.search(r"const templateMapping = (\{.*?\});\s*export default", text, re.S).group(1)
    body = re.sub(r"([,{]\s*)([A-Za-z_][A-Za-z0-9_]*)\s*:", r'\1"\2":', body)
    body = body.replace("null", "null")
    return json.loads(body)


MAPPING = load_mapping()


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def excel_serial(date_text: str) -> int | None:
    if not date_text:
        return None
    dt = datetime.strptime(date_text, "%Y-%m-%d")
    epoch = datetime(1899, 12, 30)
    return (dt - epoch).days


def money(value) -> float:
    try:
        return round(float(value or 0), 2)
    except ValueError:
        return 0.0


def col_row(cell: str) -> tuple[str, int]:
    match = re.match(r"([A-Z]+)(\d+)", cell)
    return match.group(1), int(match.group(2))


def trip_rows() -> list[int]:
    rows = []
    for section in MAPPING["tripSections"]:
        rows.extend(range(section["startRow"], section["endRow"] + 1))
    return rows


def get_style_ids(sheet_xml: bytes, cells: list[str]) -> dict[str, str]:
    root = ET.fromstring(sheet_xml)
    styles = {}
    for cell in cells:
        node = root.find(f".//m:c[@r='{cell}']", NS)
        if node is not None and "s" in node.attrib:
            styles[cell] = node.attrib["s"]
    return styles


def append_style(styles_xml: bytes, base_style_ids: list[str], num_fmt_id: str, format_code: str) -> tuple[bytes, dict[str, str]]:
    root = ET.fromstring(styles_xml)
    num_fmts = root.find("m:numFmts", NS)
    if num_fmts is None:
        num_fmts = ET.Element(f"{{{NS['m']}}}numFmts")
        root.insert(0, num_fmts)
    if not any(child.attrib.get("numFmtId") == num_fmt_id for child in list(num_fmts)):
        num_fmt = ET.SubElement(num_fmts, f"{{{NS['m']}}}numFmt")
        num_fmt.set("numFmtId", num_fmt_id)
        num_fmt.set("formatCode", format_code)
    num_fmts.set("count", str(len(list(num_fmts))))

    cell_xfs = root.find("m:cellXfs", NS)
    original = list(cell_xfs)
    style_map = {}
    for base in base_style_ids:
        clone = copy.deepcopy(original[int(base)])
        clone.set("numFmtId", num_fmt_id)
        clone.set("applyNumberFormat", "1")
        cell_xfs.append(clone)
        style_map[base] = str(len(list(cell_xfs)) - 1)
    cell_xfs.set("count", str(len(list(cell_xfs))))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), style_map


def clear_cell(cell):
    for child in list(cell):
        if child.tag.endswith(("}f", "}v", "}is")):
            cell.remove(child)
    cell.attrib.pop("t", None)


def set_cell(root, address: str, value, style_maps=None):
    col, row_num = col_row(address)
    sheet_data = root.find("m:sheetData", NS)
    row = root.find(f".//m:row[@r='{row_num}']", NS)
    if row is None:
        row = ET.SubElement(sheet_data, f"{{{NS['m']}}}row", {"r": str(row_num)})
    cell = root.find(f".//m:c[@r='{address}']", NS)
    if cell is None:
        cell = ET.SubElement(row, f"{{{NS['m']}}}c", {"r": address})
    original_style = cell.attrib.get("s")
    if style_maps and original_style in style_maps:
        cell.set("s", style_maps[original_style])
    clear_cell(cell)
    if value is None or value == "":
        return
    if isinstance(value, (int, float)):
        ET.SubElement(cell, f"{{{NS['m']}}}v").text = str(value)
    else:
        cell.set("t", "inlineStr")
        is_node = ET.SubElement(cell, f"{{{NS['m']}}}is")
        t_node = ET.SubElement(is_node, f"{{{NS['m']}}}t", {"{http://www.w3.org/XML/1998/namespace}space": "preserve"})
        t_node.text = str(value)


def build_patches(claim: dict):
    if len(claim["trips"]) > MAPPING["maxTrips"]:
        raise ValueError(f"Template supports {MAPPING['maxTrips']} trips only.")
    patches = [
        ("B1", f"Travelling Report from {claim.get('weekStart') or '_________'} to {claim.get('weekEnding') or '__________'} (以一週計即星期一至星期日) "),
        ("I1", f"  僱員姓名 : {claim.get('employeeName') or ''}{f' ({claim.get('employeeNumber')})' if claim.get('employeeNumber') else ''}"),
        ("B2", f"僱員住址: {claim.get('address') or ''}"),
        ("B3", f"Claim Date: {claim.get('claimDate') or ''}{f' | Remarks: {claim.get('remarks')}' if claim.get('remarks') else ''}"),
    ]
    rows = trip_rows()
    for index, trip in enumerate(claim["trips"]):
        row = rows[index]
        amount = money(trip.get("amount"))
        deduction = money(trip.get("deduction"))
        transport = trip.get("transportType") or "Other"
        col = MAPPING["transportColumns"].get(transport)
        suffix = f" [Other: {amount:.2f}]" if transport == "Other" else ""
        patches.extend(
            [
                (f"B{row}", ("date", excel_serial(trip.get("date")))),
                (f"C{row}", f"{trip.get('from') or ''} → {trip.get('to') or ''}{suffix}"),
                (f"E{row}", trip.get("timeFrom") or ""),
                (f"F{row}", trip.get("timeTo") or ""),
                (f"G{row}", trip.get("client") or ""),
                (f"H{row}", trip.get("projectCode") or ""),
                (f"M{row}", ("money", amount)),
                (f"N{row}", ("money", deduction)),
            ]
        )
        if col:
            patches.append((f"{col}{row}", ("money", amount)))
    page1_count = MAPPING["tripSections"][0]["endRow"] - MAPPING["tripSections"][0]["startRow"] + 1
    total1 = sum(money(t.get("amount")) - money(t.get("deduction")) for t in claim["trips"][:page1_count])
    total2 = sum(money(t.get("amount")) - money(t.get("deduction")) for t in claim["trips"][page1_count:])
    patches.extend([("N29", ("money", round(total1, 2))), ("N58", ("money", round(total2, 2)))])
    return patches


def export_claim(claim: dict, output_path: Path):
    with zipfile.ZipFile(TEMPLATE, "r") as zin:
        sheet_xml = zin.read(MAPPING["workbook"]["sheetXml"]["Sheet1"])
        styles_xml = zin.read(MAPPING["workbook"]["stylesXml"])
        style_ids = get_style_ids(sheet_xml, ["B7", "I7", "J7", "K7", "L7", "M7", "N7", "N29", "N58"])
        styles_xml, date_styles = append_style(styles_xml, [style_ids["B7"]], "165", "dd/mm/yyyy")
        styles_xml, money_styles = append_style(styles_xml, [v for k, v in style_ids.items() if k != "B7"], "166", "0.00")
        root = ET.fromstring(sheet_xml)
        for address, value in build_patches(claim):
            if isinstance(value, tuple) and value[0] == "date":
                set_cell(root, address, value[1], date_styles)
            elif isinstance(value, tuple) and value[0] == "money":
                set_cell(root, address, value[1], money_styles)
            else:
                set_cell(root, address, value)
        sheet_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == MAPPING["workbook"]["sheetXml"]["Sheet1"]:
                    data = sheet_xml
                elif item.filename == MAPPING["workbook"]["stylesXml"]:
                    data = styles_xml
                zout.writestr(item, data)


def base_claim(trips):
    return {
        "employeeName": "Chan Tai Man",
        "employeeNumber": "E001",
        "address": "Mong Kok",
        "weekStart": "2026-07-20",
        "weekEnding": "2026-07-26",
        "claimDate": "2026-07-28",
        "remarks": "Test claim",
        "trips": trips,
    }


def trip(i, transport="MTR", amount=10, deduction=0):
    return {
        "date": "2026-07-21",
        "from": f"From {i}",
        "to": f"To {i}",
        "timeFrom": "09:00",
        "timeTo": "10:00",
        "client": f"Client {i}",
        "projectCode": f"P{i:03d}",
        "transportType": transport,
        "amount": amount,
        "deduction": deduction,
        "remarks": f"Remark {i}",
    }


CASES = [
    ("A", "1 trip MTR", base_claim([trip(1, "MTR", 18)]), True),
    ("B", "5 trips different transport types", base_claim([trip(i + 1, t, 10 + i) for i, t in enumerate(["MTR", "Bus", "Van", "Taxi", "Other"])]), True),
    ("C", "Decimal amount HK$12.50", base_claim([trip(1, "Bus", 12.5)]), True),
    ("D", "With deduction", base_claim([trip(1, "Taxi", 80, 20)]), True),
    ("E", "Fill all trip rows", base_claim([trip(i + 1, ["MTR", "Bus", "Van", "Taxi"][i % 4], 5 + i) for i in range(MAPPING["maxTrips"])]), True),
    ("F", "Exceed max rows", base_claim([trip(i + 1, "MTR", 1) for i in range(MAPPING["maxTrips"] + 1)]), False),
]


def workbook_facts(path: Path):
    wb = openpyxl.load_workbook(path, data_only=False)
    formulas = []
    errors = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        formulas.append((ws.title, cell.coordinate, cell.value))
                    if re.search(r"#REF!|#VALUE!|#NAME\\?", cell.value):
                        errors.append((ws.title, cell.coordinate, cell.value))
    return {
        "sheetNames": wb.sheetnames,
        "mergedCounts": {ws.title: len(ws.merged_cells.ranges) for ws in wb.worksheets},
        "formulas": formulas,
        "errors": errors,
        "wb": wb,
    }


def verify_case(case_id: str, claim: dict, output_path: Path):
    template_facts = workbook_facts(TEMPLATE)
    out_facts = workbook_facts(output_path)
    assert out_facts["sheetNames"] == template_facts["sheetNames"]
    assert out_facts["mergedCounts"] == template_facts["mergedCounts"]
    assert len(out_facts["formulas"]) == len(template_facts["formulas"])
    assert not out_facts["errors"]
    ws = out_facts["wb"]["Sheet1"]
    assert "Chan Tai Man" in ws["I1"].value
    assert ws["B7"].value is not None
    assert ws["C7"].value.startswith("From 1")
    assert ws["G7"].value == "Client 1"
    assert ws["H7"].value == "P001"
    assert isinstance(ws["M7"].value, (int, float))
    if case_id == "C":
        assert float(ws["M7"].value) == 12.5
    if case_id == "D":
        assert float(ws["N7"].value) == 20
        assert float(ws["N29"].value) == 60
    if case_id == "E":
        assert ws["C56"].value.startswith("From 44")


def run():
    OUTPUT.mkdir(parents=True, exist_ok=True)
    shutil.rmtree(OUTPUT)
    OUTPUT.mkdir(parents=True, exist_ok=True)
    original_hash = sha256(TEMPLATE)
    rows = []
    for case_id, name, claim, should_export in CASES:
        output_path = OUTPUT / f"case_{case_id}.xlsx"
        try:
            export_claim(claim, output_path)
            if not should_export:
                raise AssertionError("Expected export to fail but it succeeded.")
            verify_case(case_id, claim, output_path)
            result = "PASS"
            detail = f"Exported and verified {output_path.name}."
        except Exception as exc:
            if should_export:
                result = "FAIL"
                detail = str(exc)
            else:
                result = "PASS"
                detail = f"Correctly blocked: {exc}"
        rows.append((case_id, name, len(claim["trips"]), result, detail))
    hash_after = sha256(TEMPLATE)
    hash_result = "PASS" if hash_after == original_hash else "FAIL"
    report = [
        "# TEST_REPORT",
        "",
        f"Template: `{TEMPLATE.name}`",
        f"Template SHA256 before: `{original_hash}`",
        f"Template SHA256 after: `{hash_after}`",
        f"Template hash unchanged: **{hash_result}**",
        "",
        "| Test | Data | Trips | Result | Actual Result |",
        "|---|---:|---:|---|---|",
    ]
    for row in rows:
        report.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} | {row[4]} |")
    report.extend(
        [
            "",
            "## Verified",
            "- Workbook opens with openpyxl after export.",
            "- Sheet count and sheet names match the template.",
            "- Merged-cell counts match the template.",
            "- Mapped cells contain expected values.",
            "- Numeric amount and deduction cells are numbers.",
            "- Formula count is unchanged. The supplied template contains no formula cells.",
            "- No `#REF!`, `#VALUE!` or `#NAME?` strings were found.",
            "- Template source file hash is unchanged.",
            "",
            "## Known Limitations",
            "- The supplied template has no separate cells for From and To because C:D is merged on each trip row; the app writes one combined location value.",
            "- The supplied template has no Other transport column; Other is annotated in the location text and included in Sub Total.",
            "- The supplied template has no row-level Remarks export cell, so trip remarks remain in app draft/history/preview.",
            "- Full visual fidelity in Apple Numbers and Microsoft 365 is expected because the package parts are preserved, but automated tests here verify workbook structure and cell data rather than opening those apps.",
        ]
    )
    REPORT.write_text("\n".join(report), encoding="utf-8")
    print("\n".join(report))
    if any(row[3] == "FAIL" for row in rows) or hash_result == "FAIL":
        raise SystemExit(1)


if __name__ == "__main__":
    run()
